# Alexa skill
"""
 
All coordinates assume a screen resolution of 1600x900, and Chrome 
maximized with the Bookmarks Toolbar enabled.
Down key has been hit 0 times to center play area in browser.
x_pad = 304
y_pad = 342
Play area =  x_pad+1, y_pad+1, 796, 825
881,494
"""
import ssl
import sys
import socket
import webbrowser
import pyscreenshot as ImageGrab
import os
import time
import re
import subprocess
import win32con
import win32api
from PIL import Image, ImageOps
from numpy import *
import turtle
import webbrowser
import speedtest
import Cited
import subprocess
import shlex  
from subprocess import Popen, PIPE, STDOUT
from xlwt import Workbook
import xlrd
import datetime
import openpyxl
#print (windows-default)
#get('windows-default')
#chrome_path="C:\Program Files (x86)\Google\Chrome\Application"

def mytime():
    time
def leftDown():
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN,0,0)
    time.sleep(.1)
    print ('left Down')
         
def leftUp():
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP,0,0)
    time.sleep(.1)
    print ('left release')
def get_cords():
    x,y = win32api.GetCursorPos()
    x = x - x_pad
    y = y - y_pad
    print (x,y)
def leftClick():
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN,0,0)
    time.sleep(.1)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP,0,0)
    print ( "Click.")
def mousePos():
    win32api.SetCursorPos((881, 494))
def screenGrab():
    im = ImageGrab.grab()
    #also can use bbox in im=... instead of calling box in
    im.show()
    im.save(os.getcwd() + '\\full_snap__' + str(int(time.time())) + '.png', 'PNG')
    return im
def __init__(self, download=0, upload=0, ping=0, server=None, client=None,
                 opener=None, secure=False):
        self.download = download
        self.upload = upload
        self.ping = ping
        if server is None:
            self.server = {}
        else:
            self.server = server
        self.client = client or {}

        self._share = None
        self.timestamp = '%sZ' % datetime.datetime.utcnow().isoformat()
        self.bytes_received = 0
        self.bytes_sent = 0

        if opener:
            self._opener = opener
        else:
            self._opener = build_opener()

        self._secure = secure

def __repr__(self):
        return repr(self.dict())

def share(self):
        """POST data to the speedtest.net API to obtain a share results
        link
        """

        if self._share:
            return self._share

        download = int(round(self.download / 1000.0, 0))
        ping = int(round(self.ping, 0))
        upload = int(round(self.upload / 1000.0, 0))
def get_simple_cmd_output(cmd, stderr=STDOUT):
    """
    Execute a simple external command and get its output.
    """
    args = shlex.split(cmd)
    return Popen(args, stdout=PIPE, stderr=stderr).communicate()[0]
 
def get_ping_time(host):
    host = host.split(':')[0]
    cmd = "fping {host} -C 3 -q".format(host=host)
    res = [float(x) for x in process.get_simple_cmd_output(cmd).strip().split(':')[-1].split() if x != '-']
    if len(res) > 0:
        return sum(res) / len(res)
    else:
        return 999999        

def main():
    speedtester = speedtest.Speedtest()
    speedtester.get_best_server()
    p = subprocess.Popen(["ping.exe","www.speedtest.net"], stdout = subprocess.PIPE)
    Get_Ping = p.communicate()[0]
    Ping = print ('Your Ping is',str (Get_Ping)[492:-7], 'Miliseconds')
    Download= speedtester.download()
    Download_Speed= round(Download/ 1000000)
    print ('Your Download Speed is',Download_Speed, 'Megabits per second')
    Upload= speedtester.upload()
    Upload_Speed= round(Upload/1000000)
    print ('Your Upload Speed is',Upload_Speed, 'Megabits per second')
    """wb = Workbook()
    sheet1 = wb.add_sheet('Network Test Results1')
    sheet1.write(1,0,datetime.datetime.now())
    sheet1.write(1,1,str (Get_Ping)[492:-7])
    sheet1.write(1,2,Download_Speed)
    sheet1.write(1,3,Upload_Speed)
    wb.save('Network Test Results.xls')"""
    book = openpyxl.load_workbook('Network Speeds.xlsx')
    sheet = book['Network Test Results1']
    r = sheet.max_row
    r= r+1
    sheet.cell(row=r,column=1,).value = now = time.ctime(int(time.time()))
    sheet.cell(row=r,column=2,).value =str (Get_Ping)[492:-7]
    sheet.cell(row=r,column=3,).value =Download_Speed
    sheet.cell(row=r,column=4,).value =Upload_Speed
    book.save('Network Speeds.xlsx')
if __name__ == '__main__':
     main()

    
    


